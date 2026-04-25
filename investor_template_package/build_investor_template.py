from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabel
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import datetime

wb = Workbook()

# ── Color Palette ──────────────────────────────────────────────────────────
DARK_NAVY   = "1B2A4A"
GOLD        = "C9A84C"
LIGHT_GOLD  = "F5E6C0"
DARK_GREEN  = "1A6B3C"
LIGHT_GREEN = "D6F0E0"
DARK_RED    = "8B0000"
LIGHT_RED   = "FFE0E0"
BLUE_HEADER = "1F4E79"
LIGHT_BLUE  = "DEEAF1"
ORANGE      = "C55A11"
LIGHT_ORANGE= "FCE4D6"
YELLOW_BG   = "FFFF00"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
MED_GRAY    = "D9D9D9"
PURPLE      = "7030A0"
LIGHT_PURPLE= "EAD1F5"
TEAL        = "008080"
LIGHT_TEAL  = "D0EBEB"

def side(style="thin", color="BFBFBF"):
    return Side(style=style, color=color)

def border(all="thin"):
    s = side(all)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = side("medium", "1B2A4A")
    return Border(left=t, right=t, top=t, bottom=t)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col_widths(ws, widths_dict):
    for col, w in widths_dict.items():
        ws.column_dimensions[col].width = w

def apply_header_row(ws, row, headers, start_col=1, bg=BLUE_HEADER, fg=WHITE):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col+i, value=h)
        c.font = font(bold=True, color=fg, size=9)
        c.fill = fill(bg)
        c.alignment = align("center")
        c.border = border("thin")

def section_title(ws, row, col, text, span=1, bg=DARK_NAVY, fg=GOLD):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, color=fg, size=11, name="Arial")
    c.fill = fill(bg)
    c.alignment = align("center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c.border = thick_border()

def data_cell(ws, row, col, value, bg=WHITE, fg="000000", bold=False,
              num_fmt=None, h_align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h_align)
    c.border = border()
    if num_fmt:
        c.number_format = num_fmt
    return c

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws_dash = wb.active
ws_dash.title = "📊 Dashboard"
ws_dash.sheet_view.showGridLines = False
ws_dash.row_dimensions[1].height = 50
set_col_widths(ws_dash, {
    "A":2,"B":18,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14,"I":14,"J":14,"K":14,"L":14
})

# Title Banner
ws_dash.merge_cells("B1:L1")
c = ws_dash["B1"]
c.value = "📈  INVESTOR PORTFOLIO COMMAND CENTER"
c.font = Font(bold=True, color=WHITE, size=20, name="Arial")
c.fill = fill(DARK_NAVY)
c.alignment = align("center")

ws_dash.merge_cells("B2:L2")
c = ws_dash["B2"]
c.value = f"Report Date: {datetime.date.today().strftime('%d %B %Y')}   |   Update your data in each sheet tab below"
c.font = font(color=GOLD, size=10, bold=True)
c.fill = fill(BLUE_HEADER)
c.alignment = align("center")
ws_dash.row_dimensions[2].height = 18

# ── KPI Cards ──
kpi_row = 4
kpis = [
    ("💰 Total Portfolio Value", "=SUM('📋 Portfolio'!H2:H21)", "₹#,##0.00", BLUE_HEADER, WHITE),
    ("📈 Total Invested", "=SUM('📋 Portfolio'!G2:G21)", "₹#,##0.00", DARK_NAVY, WHITE),
    ("💹 Overall P&L", "=SUM('📋 Portfolio'!I2:I21)", "₹#,##0.00;(₹#,##0.00);-", DARK_GREEN, WHITE),
    ("📊 Overall P&L %", "=(SUM('📋 Portfolio'!H2:H21)-SUM('📋 Portfolio'!G2:G21))/IF(SUM('📋 Portfolio'!G2:G21)=0,1,SUM('📋 Portfolio'!G2:G21))", "0.00%;(0.00%);-", GOLD, DARK_NAVY),
    ("🎯 GTT Orders Active", "=COUNTA('🎯 GTT Plan'!B2:B21)", "0", PURPLE, WHITE),
    ("🔄 Swing Trades Open", "=COUNTA('🔄 Swing Trading'!B2:B16)", "0", ORANGE, WHITE),
]

ws_dash.merge_cells("B3:L3")
c = ws_dash["B3"]
c.value = "KEY PORTFOLIO METRICS"
c.font = Font(bold=True, color=WHITE, size=10, name="Arial")
c.fill = fill(BLUE_HEADER)
c.alignment = align("center")
ws_dash.row_dimensions[3].height = 16

col_positions = ["B","C","D","E","F","G","H","I","J","K"]
kpi_cols = [("B","C"), ("D","E"), ("F","G"), ("H","I"), ("J","K"), ("L","L")]

for idx, (label, formula, fmt, bg, fg) in enumerate(kpis):
    c1 = ["B","D","F","H","J","L"][idx]
    c2 = ["C","E","G","I","K","L"][idx]
    if c1 != c2:
        ws_dash.merge_cells(f"{c1}{kpi_row}:{c2}{kpi_row}")
        ws_dash.merge_cells(f"{c1}{kpi_row+1}:{c2}{kpi_row+1}")
    c = ws_dash[f"{c1}{kpi_row}"]
    c.value = label
    c.font = Font(bold=True, color=fg, size=9, name="Arial")
    c.fill = fill(bg)
    c.alignment = align("center")
    c.border = border("medium")
    ws_dash.row_dimensions[kpi_row].height = 20

    v = ws_dash[f"{c1}{kpi_row+1}"]
    v.value = formula
    v.font = Font(bold=True, color=fg, size=14, name="Arial")
    v.fill = fill(bg)
    v.alignment = align("center")
    v.number_format = fmt
    v.border = border("medium")
    ws_dash.row_dimensions[kpi_row+1].height = 28

# ── Section Titles Row 7 ──
r = 7
ws_dash.row_dimensions[r].height = 6

# Portfolio Summary mini-table
r = 8
section_title(ws_dash, r, 2, "📋 PORTFOLIO SNAPSHOT", 5)
headers_snap = ["Stock", "Qty", "Avg Cost", "CMP", "Current Value", "P&L"]
apply_header_row(ws_dash, r+1, headers_snap, 2, BLUE_HEADER)
ws_dash.row_dimensions[r+1].height = 16

sample_stocks = [
    ("RELIANCE","50","2400.00","2520.00"),
    ("INFY","100","1500.00","1620.00"),
    ("TCS","25","3400.00","3600.00"),
    ("HDFCBANK","75","1650.00","1580.00"),
    ("WIPRO","200","450.00","430.00"),
]
for i, (st, qty, avg, cmp) in enumerate(sample_stocks):
    row = r+2+i
    bg_ = LIGHT_GRAY if i%2==0 else WHITE
    for j, v in enumerate([st, qty, avg, cmp]):
        data_cell(ws_dash, row, 2+j, v, bg_, num_fmt="0" if j<2 else "₹#,##0.00")
    # Current Value
    cv = ws_dash.cell(row=row, column=6, value=f"=C{row}*D{row}")
    cv.number_format="₹#,##0.00"; cv.fill=fill(bg_); cv.alignment=align("center"); cv.border=border()
    # P&L
    pl = ws_dash.cell(row=row, column=7, value=f"=F{row}-(C{row}*E{row})")
    pl.number_format="₹#,##0.00;(₹#,##0.00);-"; pl.fill=fill(bg_); pl.alignment=align("center"); pl.border=border()

# Conditional format P&L column green/red
from openpyxl.formatting.rule import CellIsRule
green_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
red_fill   = PatternFill(start_color=LIGHT_RED,   end_color=LIGHT_RED,   fill_type="solid")
ws_dash.conditional_formatting.add(f"G{r+2}:G{r+6}", CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill))
ws_dash.conditional_formatting.add(f"G{r+2}:G{r+6}", CellIsRule(operator="lessThan",    formula=["0"], fill=red_fill))

# Right side – Buy suggestions summary
section_title(ws_dash, 8, 8, "💡 TOP BUY SUGGESTIONS TODAY", 4)
apply_header_row(ws_dash, 9, ["Stock", "Rating", "Target", "Upside%"], 8, DARK_GREEN)
buys = [("HCLTECH","⭐⭐⭐⭐⭐","1400","18%"),("SUNPHARMA","⭐⭐⭐⭐","1050","12%"),
        ("POWERGRID","⭐⭐⭐⭐","280","15%"),("COALINDIA","⭐⭐⭐","270","10%"),("LT","⭐⭐⭐⭐⭐","3800","20%")]
for i, (st, rat, tgt, up) in enumerate(buys):
    bg_ = LIGHT_GREEN if i%2==0 else WHITE
    for j, v in enumerate([st, rat, tgt, up]):
        data_cell(ws_dash, 10+i, 8+j, v, bg_)

# Buyback Opportunities summary
r2 = 16
section_title(ws_dash, r2, 2, "🔁 BUYBACK OPPORTUNITIES", 5)
apply_header_row(ws_dash, r2+1, ["Company", "Buyback Price", "CMP", "Premium%", "Status"], 2, PURPLE)
buybacks = [("INFY","₹1,800","₹1,620","11.1%","🟢 Open"),("TCS","₹4,200","₹3,600","16.7%","🟢 Open"),
            ("HCL TECH","₹1,500","₹1,185","26.6%","🟡 Expected"),("WIPRO","₹500","₹430","16.3%","🟢 Open")]
for i, row_data in enumerate(buybacks):
    bg_ = LIGHT_PURPLE if i%2==0 else WHITE
    for j, v in enumerate(row_data):
        data_cell(ws_dash, r2+2+i, 2+j, v, bg_)

# GTT mini
section_title(ws_dash, r2, 8, "🎯 ACTIVE GTT ORDERS", 4)
apply_header_row(ws_dash, r2+1, ["Stock", "Trigger", "Type", "Status"], 8, PURPLE)
gtts = [("RELIANCE","2300","Buy on dip","🟢 Active"),("INFY","1550","Stop Loss","🟡 Pending"),
        ("HDFCBANK","1500","Buy Target","🟢 Active"),("TCS","3300","Profit Book","🔵 Set")]
for i, row_data in enumerate(gtts):
    bg_ = LIGHT_PURPLE if i%2==0 else WHITE
    for j, v in enumerate(row_data):
        data_cell(ws_dash, r2+2+i, 8+j, v, bg_)

# Swing Trade mini
r3 = 22
section_title(ws_dash, r3, 2, "🔄 SWING TRADE WATCHLIST", 11)
apply_header_row(ws_dash, r3+1, ["Stock","Setup","Entry Zone","Target","Stop Loss","Reward:Risk","Days","Confidence","Action"], 2, ORANGE)
swings = [
    ("INFY","Buyback+Dip","1600-1650","1780","1550","3.6:1","5-8","🔥 High","BUY"),
    ("WIPRO","Buyback Rally","425-435","490","410","2.8:1","7-10","🔥 High","BUY"),
    ("TCS","Value+Buyback","3550-3650","3900","3450","2.5:1","10-15","✅ Med","WATCH"),
    ("HCL TECH","Pre-Buyback","1150-1200","1380","1100","3.2:1","8-12","🔥 High","BUY"),
]
for i, row_data in enumerate(swings):
    bg_ = LIGHT_ORANGE if i%2==0 else WHITE
    fg_ = DARK_RED if row_data[-1]=="WATCH" else DARK_GREEN
    for j, v in enumerate(row_data):
        c = data_cell(ws_dash, r3+2+i, 2+j, v, bg_)
        if j == len(row_data)-1:
            c.font = Font(bold=True, color=fg_, size=9, name="Arial")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — PORTFOLIO
# ══════════════════════════════════════════════════════════════════════════════
ws_port = wb.create_sheet("📋 Portfolio")
ws_port.sheet_view.showGridLines = False
set_col_widths(ws_port, {"A":2,"B":20,"C":12,"D":12,"E":12,"F":14,"G":16,"H":16,"I":14,"J":12,"K":14,"L":14,"M":14})

ws_port.merge_cells("B1:M1")
c = ws_port["B1"]
c.value = "📋  MY INVESTMENT PORTFOLIO"
c.font = Font(bold=True, color=WHITE, size=16, name="Arial")
c.fill = fill(DARK_NAVY)
c.alignment = align("center")
ws_port.row_dimensions[1].height = 36

headers_port = ["Stock / Company","Sector","Exchange","Buy Date","Qty","Avg Buy Price (₹)",
                "Total Invested (₹)","Current Value (₹)","P&L (₹)","P&L %","Weight %","52W High","52W Low"]
apply_header_row(ws_port, 2, headers_port, 2, BLUE_HEADER)
ws_port.row_dimensions[2].height = 32

sample_port = [
    ("RELIANCE","Energy","NSE","01-Jan-2024","50","2400","","","","","","2800","2100"),
    ("INFY","IT","NSE","15-Feb-2024","100","1500","","","","","","1900","1350"),
    ("TCS","IT","NSE","10-Mar-2024","25","3400","","","","","","4000","3100"),
    ("HDFCBANK","Banking","NSE","05-Apr-2024","75","1650","","","","","","1900","1400"),
    ("WIPRO","IT","NSE","20-May-2024","200","450","","","","","","570","380"),
    ("SUNPHARMA","Pharma","NSE","01-Jun-2024","80","900","","","","","","1100","780"),
    ("HCLTECH","IT","NSE","15-Jul-2024","60","1200","","","","","","1500","1050"),
    ("POWERGRID","Utilities","NSE","01-Aug-2024","300","240","","","","","","310","200"),
    ("COALINDIA","Mining","NSE","10-Sep-2024","150","245","","","","","","295","210"),
    ("LT","Infra","NSE","15-Oct-2024","30","3200","","","","","","3900","2800"),
]

for i, row_data in enumerate(sample_port):
    r = 3+i
    bg_ = LIGHT_BLUE if i%2==0 else WHITE
    # Static cols
    for j in range(5):
        data_cell(ws_port, r, 2+j, row_data[j], bg_, h_align="left" if j==0 else "center")
    # Qty (blue input)
    c = ws_port.cell(row=r, column=6, value=int(row_data[4]))
    c.font = Font(color="0000FF", size=9, name="Arial"); c.fill=fill(bg_)
    c.alignment=align("center"); c.border=border()
    # Avg Buy Price (blue input)
    c = ws_port.cell(row=r, column=7, value=float(row_data[5]))
    c.font = Font(color="0000FF", size=9, name="Arial"); c.fill=fill(bg_)
    c.number_format="₹#,##0.00"; c.alignment=align("center"); c.border=border()
    # Current Value placeholder (user fills CMP, formula computes)
    cmp_col = 8  # H = CMP (user input, blue)
    c = ws_port.cell(row=r, column=cmp_col, value=0.00)
    c.font=Font(color="0000FF", size=9, name="Arial"); c.fill=fill(LIGHT_GOLD)
    c.number_format="₹#,##0.00"; c.alignment=align("center"); c.border=border()
    ws_port.cell(row=2, column=cmp_col).value = "CMP (₹) [INPUT]"
    # Total Invested
    ti = ws_port.cell(row=r, column=9, value=f"=F{r}*G{r}")
    ti.number_format="₹#,##0.00"; ti.fill=fill(bg_); ti.alignment=align("center"); ti.border=border()
    # Current Value = Qty * CMP
    cv = ws_port.cell(row=r, column=10, value=f"=F{r}*H{r}")
    cv.number_format="₹#,##0.00"; cv.fill=fill(bg_); cv.alignment=align("center"); cv.border=border()
    # P&L
    pl = ws_port.cell(row=r, column=11, value=f"=J{r}-I{r}")
    pl.number_format="₹#,##0.00;(₹#,##0.00);-"; pl.fill=fill(bg_); pl.alignment=align("center"); pl.border=border()
    # P&L%
    plp = ws_port.cell(row=r, column=12, value=f"=IF(I{r}=0,0,(J{r}-I{r})/I{r})")
    plp.number_format="0.00%;(0.00%);-"; plp.fill=fill(bg_); plp.alignment=align("center"); plp.border=border()
    # Weight%
    wt = ws_port.cell(row=r, column=13, value=f"=IF(SUM(J$3:J$22)=0,0,J{r}/SUM(J$3:J$22))")
    wt.number_format="0.00%"; wt.fill=fill(bg_); wt.alignment=align("center"); wt.border=border()
    # 52W High
    data_cell(ws_port, r, 14, float(row_data[11]), bg_, num_fmt="₹#,##0.00")
    # 52W Low
    data_cell(ws_port, r, 15, float(row_data[12]), bg_, num_fmt="₹#,##0.00")

# Fix header row
apply_header_row(ws_port, 2,
    ["Stock / Company","Sector","Exchange","Buy Date","Qty","Avg Buy Price (₹)",
     "CMP (₹) [INPUT]","Total Invested (₹)","Current Value (₹)","P&L (₹)","P&L %","Weight %","52W High","52W Low"],
    2, BLUE_HEADER)
ws_port.row_dimensions[2].height = 32

# Totals Row
tr = 13
ws_port.cell(row=tr, column=2, value="📊 TOTAL").font = Font(bold=True, size=10)
ws_port.cell(row=tr, column=2).fill=fill(DARK_NAVY)
ws_port.cell(row=tr, column=2).font=Font(bold=True, color=WHITE, size=10, name="Arial")
ws_port.cell(row=tr, column=2).alignment=align("center")
for c_idx, formula, fmt in [
    (9, "=SUM(I3:I12)", "₹#,##0.00"),
    (10, "=SUM(J3:J12)", "₹#,##0.00"),
    (11, "=SUM(K3:K12)", "₹#,##0.00;(₹#,##0.00);-"),
    (12, "=IF(I13=0,0,(J13-I13)/I13)", "0.00%;(0.00%);-"),
]:
    c = ws_port.cell(row=tr, column=c_idx, value=formula)
    c.font=Font(bold=True, size=10, name="Arial"); c.fill=fill(LIGHT_GOLD)
    c.number_format=fmt; c.alignment=align("center"); c.border=thick_border()

# CF for P&L%
ws_port.conditional_formatting.add("L3:L12", CellIsRule("greaterThan", ["0"], fill=PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")))
ws_port.conditional_formatting.add("L3:L12", CellIsRule("lessThan",    ["0"], fill=PatternFill(start_color=LIGHT_RED,   end_color=LIGHT_RED,   fill_type="solid")))

# Note row
note = ws_port.cell(row=15, column=2, value="📝 NOTE: Blue cells = User Input.  Yellow highlighted cell = CMP (enter today's price).  All calculations auto-update.")
note.font=Font(italic=True, color=DARK_NAVY, size=9, name="Arial")
ws_port.merge_cells("B15:M15")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — BUY SUGGESTIONS
# ══════════════════════════════════════════════════════════════════════════════
ws_buy = wb.create_sheet("💡 Buy Suggestions")
ws_buy.sheet_view.showGridLines = False
set_col_widths(ws_buy, {"A":2,"B":20,"C":12,"D":12,"E":12,"F":14,"G":14,"H":14,"I":16,"J":16,"K":16,"L":16,"M":16})

ws_buy.merge_cells("B1:M1")
c=ws_buy["B1"]; c.value="💡  BUY SUGGESTIONS — FUNDAMENTAL & TECHNICAL ANALYSIS"
c.font=Font(bold=True, color=WHITE, size=15, name="Arial"); c.fill=fill(DARK_GREEN)
c.alignment=align("center"); ws_buy.row_dimensions[1].height=36

# Legend
ws_buy.merge_cells("B2:M2")
c=ws_buy["B2"]; c.value="⭐⭐⭐⭐⭐ = Strong Buy  |  ⭐⭐⭐⭐ = Buy  |  ⭐⭐⭐ = Accumulate  |  🔴 = Avoid/Wait"
c.font=Font(italic=True, color=DARK_NAVY, size=9, name="Arial"); c.fill=fill(LIGHT_GREEN); c.alignment=align("center")

apply_header_row(ws_buy, 3,
    ["Company","Sector","CMP (₹)","Buy Zone","Target 1","Target 2","Stop Loss",
     "Upside %","P/E Ratio","EPS Growth %","D/E Ratio","Rating","Key Reason"],
    2, DARK_GREEN)
ws_buy.row_dimensions[3].height=30

buy_data = [
    ("HCLTECH","IT","1185","1150-1200","1380","1500","1080","18-26%","22","15%","0.1","⭐⭐⭐⭐⭐","Buyback + Strong Q Results + IT Demand"),
    ("SUNPHARMA","Pharma","925","900-950","1050","1200","850","14-30%","28","18%","0.2","⭐⭐⭐⭐","US Generic Recovery + R&D Pipeline"),
    ("POWERGRID","Utilities","240","235-245","280","310","215","17-29%","18","12%","1.8","⭐⭐⭐⭐","Govt Capex + Dividend Yield 5.2%"),
    ("COALINDIA","Mining","245","238-250","270","295","220","10-20%","7","8%","0.0","⭐⭐⭐⭐","Near-Zero Debt + 8% Dividend Yield"),
    ("LT","Infra","3200","3100-3300","3800","4200","2900","19-31%","32","20%","1.5","⭐⭐⭐⭐⭐","Order Book +25% YoY + Govt Projects"),
    ("RELIANCE","Energy","2520","2400-2550","2900","3200","2250","15-27%","26","14%","0.4","⭐⭐⭐⭐","Jio + Retail + O2C Diversification"),
    ("AXISBANK","Banking","1050","1020-1070","1220","1350","970","16-29%","14","22%","0.8","⭐⭐⭐⭐","NIM Expansion + Low NPA Trend"),
    ("BAJFINANCE","NBFC","6800","6600-7000","7800","8500","6200","15-25%","30","28%","3.2","⭐⭐⭐⭐","AUM Growth + New Products"),
    ("TITAN","Consumer","3200","3100-3300","3700","4000","2950","16-25%","72","20%","0.1","⭐⭐⭐","Jewellery + Watches Market Leader"),
    ("NTPC","Energy","350","340-360","410","450","320","17-29%","15","10%","2.0","⭐⭐⭐⭐","Green Energy Expansion + Dividend"),
]

for i, row_data in enumerate(buy_data):
    r=4+i
    bg_=LIGHT_GREEN if i%2==0 else WHITE
    stars = row_data[11]
    if "⭐⭐⭐⭐⭐" in stars: bg2=LIGHT_GREEN
    elif "⭐⭐⭐⭐" in stars: bg2=WHITE
    else: bg2=LIGHT_ORANGE
    for j, v in enumerate(row_data):
        bg_use = bg2 if j==11 else bg_
        c=data_cell(ws_buy, r, 2+j, v, bg_use, h_align="left" if j in [0,12] else "center")
        if j==12:
            c.alignment=align("left",wrap=True); ws_buy.row_dimensions[r].height=24

# Note section
r_note = 15
ws_buy.merge_cells(f"B{r_note}:M{r_note}")
c=ws_buy.cell(row=r_note, column=2, value="KEY ANALYSIS POINTS — WHAT TO CHECK BEFORE BUYING")
c.font=Font(bold=True, color=WHITE, size=10, name="Arial"); c.fill=fill(DARK_NAVY); c.alignment=align("center")

points = [
    ("1️⃣ Fundamentals", "P/E < Sector Average | EPS Growth > 15% | Debt/Equity < 1 | ROE > 15% | Strong Cash Flow"),
    ("2️⃣ Technicals",   "Price above 200 DMA | RSI 40–60 (not overbought) | MACD Crossover | Volume spike on buy"),
    ("3️⃣ Valuation",    "Check P/B Ratio vs Peers | Dividend Yield > 2% is bonus | Compare EV/EBITDA"),
    ("4️⃣ Catalysts",    "Upcoming earnings | Buyback announcement | New product/order | Sector tailwind"),
    ("5️⃣ Risk Check",   "Promoter Pledge < 10% | No major litigation | No auditor concern | FII/DII trend positive"),
    ("6️⃣ Entry Strategy","Buy in 2–3 tranches | First buy at lower zone | Average on 5% dip | Set GTT immediately"),
    ("7️⃣ Exit Plan",    "Book 50% at Target 1 | Let rest run to Target 2 | Trail stop loss | Never average loss"),
]
for i, (cat, detail) in enumerate(points):
    r=r_note+1+i
    ws_buy.row_dimensions[r].height=20
    c1=ws_buy.cell(row=r, column=2, value=cat)
    c1.font=Font(bold=True, color=WHITE, size=9, name="Arial"); c1.fill=fill(BLUE_HEADER)
    c1.alignment=align("left"); c1.border=border()
    ws_buy.merge_cells(f"C{r}:M{r}")
    c2=ws_buy.cell(row=r, column=3, value=detail)
    c2.font=font(size=9); c2.fill=fill(LIGHT_BLUE if i%2==0 else WHITE)
    c2.alignment=align("left"); c2.border=border()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — BUYBACK OPPORTUNITIES
# ══════════════════════════════════════════════════════════════════════════════
ws_bb = wb.create_sheet("🔁 Buyback Opps")
ws_bb.sheet_view.showGridLines = False
set_col_widths(ws_bb, {"A":2,"B":22,"C":12,"D":14,"E":14,"F":12,"G":14,"H":12,"I":14,"J":14,"K":20})

ws_bb.merge_cells("B1:K1")
c=ws_bb["B1"]; c.value="🔁  SHARE BUYBACK OPPORTUNITIES — DEEP VALUE TRACKER"
c.font=Font(bold=True, color=WHITE, size=15, name="Arial"); c.fill=fill(PURPLE)
c.alignment=align("center"); ws_bb.row_dimensions[1].height=36

# What is Buyback box
ws_bb.merge_cells("B2:K2")
c=ws_bb["B2"]; c.value="💡 WHAT IS A BUYBACK? Company repurchases its own shares → reduces supply → EPS increases → Share price rises"
c.font=Font(bold=True, italic=True, color=DARK_NAVY, size=9, name="Arial"); c.fill=fill(LIGHT_PURPLE)
c.alignment=align("center"); ws_bb.row_dimensions[2].height=16

apply_header_row(ws_bb, 3,
    ["Company","CMP (₹)","Buyback Price (₹)","Premium to CMP%","Buyback Size (₹Cr)","Acceptance Ratio%","Record Date","Status","Expected Gain%","Action Needed"],
    2, PURPLE)
ws_bb.row_dimensions[3].height=30

bb_data=[
    ("INFOSYS","1620","1800","=((D4-C4)/C4)","9300","25%","15-Jun-2025","🟢 OPEN","=((D4-C4)/C4)","TENDER NOW"),
    ("TCS","3600","4200","=((D5-C5)/C5)","17000","15%","30-Jun-2025","🟢 OPEN","=((D5-C5)/C5)","TENDER NOW"),
    ("WIPRO","430","500","=((D6-C6)/C6)","2500","30%","15-Jul-2025","🟢 OPEN","=((D6-C6)/C6)","TENDER NOW"),
    ("HCL TECH","1185","1500","=((D7-C7)/C7)","4500","20%","31-Aug-2025","🟡 EXPECTED","=((D7-C7)/C7)","ACCUMULATE"),
    ("MPHASIS","2100","2600","=((D8-C8)/C8)","1200","35%","30-Sep-2025","🟡 EXPECTED","=((D8-C8)/C8)","WATCH"),
]
for i, (name, cmp, bbp, prem, size, acc, date, status, gain, action) in enumerate(bb_data):
    r=4+i; bg_=LIGHT_PURPLE if i%2==0 else WHITE
    c_vals=[(name,"left"),(cmp,"center"),(bbp,"center"),(prem,"center"),
            (size,"center"),(acc,"center"),(date,"center"),(status,"center"),(gain,"center"),(action,"center")]
    for j,(v,ha) in enumerate(c_vals):
        c=ws_bb.cell(row=r, column=2+j, value=float(v) if j in [1,2] else v)
        c.fill=fill(bg_); c.alignment=align(ha); c.border=border()
        if j==3 or j==8:
            c.number_format="0.00%"
            c.font=Font(bold=True, color=DARK_GREEN, size=9, name="Arial")
        elif j in [1,2]:
            c.number_format="₹#,##0.00"
            c.font=Font(color="0000FF", size=9, name="Arial")
        elif j==9:
            fg_=DARK_GREEN if "NOW" in v else (ORANGE if "ACCUMULATE" in v else DARK_NAVY)
            c.font=Font(bold=True, color=fg_, size=9, name="Arial")
        else:
            c.font=font(size=9)

# How Buyback Arbitrage Works
r_strat = 10
section_title(ws_bb, r_strat, 2, "📚 BUYBACK ARBITRAGE — HOW IT WORKS (Step by Step)", 10)

steps = [
    ("STEP 1: Identify","Company announces buyback at PREMIUM to current market price (CMP). Example: CMP=₹1620, Buyback=₹1800"),
    ("STEP 2: Calculate","Premium % = (Buyback Price - CMP)/CMP × 100. Here: (1800-1620)/1620 × 100 = 11.1% above market"),
    ("STEP 3: Buy Shares","Buy shares in open market BEFORE record date. More shares = more tender opportunity"),
    ("STEP 4: Tender Shares","Submit shares in buyback tender offer via your broker before deadline"),
    ("STEP 5: Acceptance","Company accepts based on ratio. E.g., 25% ratio = 25 shares accepted out of 100 tendered"),
    ("STEP 6: Profit","Accepted shares sold at buyback price (₹1800) vs cost (₹1620) = ₹180/share profit"),
    ("STEP 7: Unaccepted","Remaining shares stay in your demat. Continue holding or sell in market"),
    ("⚠️ Risk Note","Acceptance ratio may be low. Price may fall post-buyback. Tax: 20% STCG if held <12 months"),
]
for i, (step, desc) in enumerate(steps):
    r=r_strat+1+i; ws_bb.row_dimensions[r].height=22
    bg1 = PURPLE if "STEP" in step else DARK_RED
    c1=ws_bb.cell(row=r, column=2, value=step)
    c1.font=Font(bold=True, color=WHITE, size=9, name="Arial"); c1.fill=fill(bg1); c1.alignment=align("center"); c1.border=border()
    ws_bb.merge_cells(f"C{r}:K{r}")
    c2=ws_bb.cell(row=r, column=3, value=desc)
    c2.font=font(size=9); c2.fill=fill(LIGHT_PURPLE if i%2==0 else WHITE); c2.alignment=align("left"); c2.border=border()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — GTT PLAN
# ══════════════════════════════════════════════════════════════════════════════
ws_gtt = wb.create_sheet("🎯 GTT Plan")
ws_gtt.sheet_view.showGridLines = False
set_col_widths(ws_gtt, {"A":2,"B":20,"C":12,"D":14,"E":14,"F":14,"G":14,"H":14,"I":14,"J":14,"K":20})

ws_gtt.merge_cells("B1:K1")
c=ws_gtt["B1"]; c.value="🎯  GTT (GOOD TILL TRIGGERED) ORDER PLANNER"
c.font=Font(bold=True, color=WHITE, size=15, name="Arial"); c.fill=fill(DARK_NAVY)
c.alignment=align("center"); ws_gtt.row_dimensions[1].height=36

ws_gtt.merge_cells("B2:K2")
c=ws_gtt["B2"]; c.value="💡 GTT = Set it and Forget it! Triggers automatically when price hits your level. Available on Zerodha Kite, Groww, Upstox, Angel One"
c.font=Font(italic=True, color=DARK_NAVY, size=9, name="Arial"); c.fill=fill(LIGHT_GOLD); c.alignment=align("center")

# GTT Types explanation
for col, (title, desc, bg) in enumerate([
    ("🟢 GTT BUY","Triggers a BUY when price FALLS to your target entry. Perfect for buying dips automatically.", DARK_GREEN),
    ("🔴 GTT SELL / STOP","Triggers a SELL when price FALLS below stop loss. Protects capital automatically.", DARK_RED),
    ("🔵 GTT PROFIT BOOK","Triggers a SELL when price RISES to target. Lock in profits without monitoring.", BLUE_HEADER),
], 1):
    ws_gtt.merge_cells(f"{get_column_letter(1+col*3)}3:{get_column_letter(3+col*3)}3")
    ws_gtt.merge_cells(f"{get_column_letter(1+col*3)}4:{get_column_letter(3+col*3)}4")
    c1=ws_gtt.cell(row=3, column=1+col*3, value=title)
    c1.font=Font(bold=True, color=WHITE, size=9, name="Arial"); c1.fill=fill(bg); c1.alignment=align("center"); c1.border=border()
    ws_gtt.row_dimensions[3].height=20
    c2=ws_gtt.cell(row=4, column=1+col*3, value=desc)
    c2.font=font(size=8); c2.fill=fill(LIGHT_BLUE); c2.alignment=align("center",wrap=True); c2.border=border()
    ws_gtt.row_dimensions[4].height=30

apply_header_row(ws_gtt, 5,
    ["Stock","CMP (₹)","GTT Type","Trigger Price (₹)","Qty","Order Value (₹)","Stop Loss","Target","R:R Ratio","Broker","Notes / Status"],
    2, DARK_NAVY)
ws_gtt.row_dimensions[5].height=28

gtt_data=[
    ("RELIANCE","2520","🟢 GTT BUY","2300","25","","2150","2800","","Zerodha Kite","Buy on 8.7% dip — strong support"),
    ("INFY","1620","🔴 GTT STOP LOSS","1530","100","","1530","1850","","Groww","Protect 5.5% downside"),
    ("TCS","3600","🔵 GTT PROFIT BOOK","3900","25","","3300","3900","","Zerodha Kite","Book at ₹3900 resistance"),
    ("HDFCBANK","1580","🟢 GTT BUY","1480","50","","1380","1800","","Upstox","Buy at 52W support zone"),
    ("WIPRO","430","🟢 GTT BUY","400","200","","375","500","","Angel One","Pre-buyback accumulation"),
    ("HCLTECH","1185","🟢 GTT BUY","1120","60","","1050","1400","","Zerodha Kite","Buyback expected — buy dip"),
    ("COALINDIA","245","🟢 GTT BUY","225","300","","210","275","","Groww","High dividend yield stock"),
    ("SUNPHARMA","925","🔵 GTT PROFIT BOOK","1050","80","","850","1050","","Zerodha Kite","Exit at target — 13.5% profit"),
]
for i, row_data in enumerate(gtt_data):
    r=6+i; bg_=LIGHT_GRAY if i%2==0 else WHITE
    gtt_type=row_data[2]
    if "BUY" in gtt_type: type_bg=LIGHT_GREEN
    elif "STOP" in gtt_type: type_bg=LIGHT_RED
    else: type_bg=LIGHT_BLUE
    # Write static fields
    static_idx = [0,1,2,3,4,6,7,9,10]
    col_map    = [2,3,4,5,6,8,9,11,12]
    for sidx, cidx in zip(static_idx, col_map):
        v = row_data[sidx]
        c=ws_gtt.cell(row=r, column=cidx, value=float(v) if cidx in [3,5,6,8,9] and str(v).replace('.','').isdigit() else v)
        bg_use=type_bg if cidx==4 else bg_
        c.fill=fill(bg_use); c.alignment=align("left" if cidx in [2,12] else "center"); c.border=border()
        if cidx in [3,5,8,9]: c.number_format="₹#,##0.00"; c.font=Font(color="0000FF",size=9,name="Arial")
        elif cidx==4:
            fg_=DARK_GREEN if "BUY" in v else (DARK_RED if "STOP" in v else BLUE_HEADER)
            c.font=Font(bold=True,color=fg_,size=9,name="Arial")
        elif cidx==6: c.font=font(size=9)
        else: c.font=font(size=9)
    # Order Value formula: Trigger(E) * Qty(F)
    ov=ws_gtt.cell(row=r, column=7, value=f"=E{r}*F{r}")
    ov.number_format="₹#,##0.00"; ov.fill=fill(bg_); ov.alignment=align("center"); ov.border=border(); ov.font=font(size=9)
    # R:R formula: (Target(I) - Trigger(E)) / (Trigger(E) - StopLoss(H))
    rr=ws_gtt.cell(row=r, column=10, value=f"=IF((E{r}-H{r})=0,0,ROUND((I{r}-E{r})/(E{r}-H{r}),1))")
    rr.number_format="0.0x"; rr.fill=fill(bg_); rr.alignment=align("center"); rr.border=border()
    rr.font=Font(bold=True,color=DARK_GREEN,size=9,name="Arial")
    ws_gtt.row_dimensions[r].height=20

# GTT Rules
r_rules = 15
section_title(ws_gtt, r_rules, 2, "📋 GTT RULES & BEST PRACTICES", 10)
rules=[
    "✅ Always set GTT immediately after buying — don't rely on memory to set stop loss",
    "✅ For every ₹1 risk (stop loss), ensure at least ₹2 reward (target) — minimum 1:2 R:R ratio",
    "✅ GTT Buy Orders: Place at strong support levels, 200 DMA, or buyback accumulation zones",
    "✅ GTT Sell Orders: Place stop at recent swing low — never more than 7-8% below entry",
    "✅ GTT Profit Book: Place first target at resistance; let remaining run with trailing stop",
    "⚠️ GTT orders expire after 1 year on most brokers — renew them periodically",
    "⚠️ GTT does not guarantee execution at exact price in fast-moving markets (gap risk)",
    "💡 TWO-LEG GTT: Set both stop loss AND target simultaneously for full automation",
]
for i, rule in enumerate(rules):
    r=r_rules+1+i
    ws_gtt.merge_cells(f"B{r}:K{r}")
    c=ws_gtt.cell(row=r,column=2,value=rule)
    c.font=Font(bold="✅" in rule or "💡" in rule, color=DARK_NAVY, size=9, name="Arial")
    c.fill=fill(LIGHT_GREEN if "✅" in rule else (LIGHT_RED if "⚠️" in rule else LIGHT_GOLD))
    c.alignment=align("left"); c.border=border(); ws_gtt.row_dimensions[r].height=18

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — SWING TRADING
# ══════════════════════════════════════════════════════════════════════════════
ws_sw = wb.create_sheet("🔄 Swing Trading")
ws_sw.sheet_view.showGridLines = False
set_col_widths(ws_sw, {"A":2,"B":20,"C":13,"D":12,"E":14,"F":14,"G":14,"H":12,"I":12,"J":12,"K":12,"L":16,"M":16})

ws_sw.merge_cells("B1:M1")
c=ws_sw["B1"]; c.value="🔄  SWING TRADING — BUYBACK & DEEP VALUE OPPORTUNITIES"
c.font=Font(bold=True,color=WHITE,size=15,name="Arial"); c.fill=fill(ORANGE)
c.alignment=align("center"); ws_sw.row_dimensions[1].height=36

ws_sw.merge_cells("B2:M2")
c=ws_sw["B2"]; c.value="Strategy: Buy deep-value stocks near buyback price support | Hold 5-15 days | Target 8-15% move | Set GTT stop immediately"
c.font=Font(italic=True,color=DARK_NAVY,size=9,name="Arial"); c.fill=fill(LIGHT_ORANGE); c.alignment=align("center")

apply_header_row(ws_sw, 3,
    ["Stock","Setup Type","Entry Zone (₹)","Target 1 (₹)","Target 2 (₹)","Stop Loss (₹)",
     "Holding Days","R:R Ratio","Confidence","Capital (₹)","P&L If T1 Hit (₹)","P&L If T2 Hit (₹)","Status"],
    2, ORANGE)
ws_sw.row_dimensions[3].height=30

sw_data=[
    ("INFY","Buyback + Dip","1600","1720","1780","1550","5-8","=((E4-C4)/(C4-F4))","🔥 HIGH","50000","=(E4-C4)/(C4)*J4","=(F4 row error fix)","🟢 ACTIVE"),
    ("WIPRO","Pre-Buyback","425","470","495","405","7-10","=((E5-C5)/(C5-F5))","🔥 HIGH","40000","","","🟢 ACTIVE"),
    ("HCL TECH","Buyback Play","1150","1300","1380","1100","8-12","=((E6-C6)/(C6-F6))","🔥 HIGH","60000","","","🟡 WATCH"),
    ("TCS","Value Rebound","3550","3800","3900","3400","10-15","=((E7-C7)/(C7-F7))","✅ MED","80000","","","🟡 WATCH"),
    ("MPHASIS","Deep Value","2050","2300","2500","1950","10-14","=((E8-C8)/(C8-F8))","✅ MED","50000","","","🟡 WATCH"),
    ("BAJFINANCE","Oversold","6600","7200","7600","6200","12-18","=((E9-C9)/(C9-F9))","✅ MED","1,00,000","","","⚪ SETUP"),
    ("HDFCBANK","Support Buy","1500","1650","1720","1440","8-12","=((E10-C10)/(C10-F10))","🔥 HIGH","75000","","","⚪ SETUP"),
    ("SUNPHARMA","Breakout","920","1020","1080","880","7-10","=((E11-C11)/(C11-F11))","✅ MED","45000","","","⚪ SETUP"),
]

for i, row_data in enumerate(sw_data):
    r=4+i; bg_=LIGHT_ORANGE if i%2==0 else WHITE
    conf=row_data[8]; stat=row_data[12]
    stat_bg=LIGHT_GREEN if "ACTIVE" in stat else (LIGHT_GOLD if "WATCH" in stat else LIGHT_GRAY)
    
    vals=[row_data[0],row_data[1],
          float(row_data[2].replace(",","")) if row_data[2].replace(",","").replace(".","").isdigit() else row_data[2],
          float(row_data[3]),float(row_data[4]),float(row_data[5]),
          row_data[6], row_data[7], conf,
          float(row_data[9].replace(",","")) if row_data[9].replace(",","").isdigit() else row_data[9],
          f"=ROUND(({get_column_letter(6)}{r}-{get_column_letter(4)}{r})/{get_column_letter(4)}{r}*{get_column_letter(11)}{r},0)" if i>0 else "=ROUND((E4-C4)/C4*J4,0)",
          f"=ROUND(({get_column_letter(7)}{r}-{get_column_letter(4)}{r})/{get_column_letter(4)}{r}*{get_column_letter(11)}{r},0)" if i>0 else "=ROUND((F4-C4)/C4*J4,0)",
          stat]
    
    # Rebuild proper P&L formulas - columns: B=stock(2),C=setup(3),D=entry(4),E=t1(5),F=t2(6),G=stop(7),H=days(8),I=rr(9),J=conf(10),K=capital(11),L=pnl_t1(12),M=pnl_t2(13),N=status(14)
    pnl_t1 = f"=IFERROR(ROUND((E{r}-D{r})/D{r}*K{r},0),0)"
    pnl_t2 = f"=IFERROR(ROUND((F{r}-D{r})/D{r}*K{r},0),0)"
    rr_formula = f"=IFERROR(ROUND((E{r}-D{r})/(D{r}-G{r}),1),0)"
    
    entries=[
        (row_data[0],"left",bg_,None,"000000",False),
        (row_data[1],"center",bg_,None,"000000",False),
        (float(row_data[2]),"center",bg_,"₹#,##0.00","0000FF",False),
        (float(row_data[3]),"center",bg_,"₹#,##0.00","0000FF",False),
        (float(row_data[4]),"center",bg_,"₹#,##0.00","0000FF",False),
        (float(row_data[5]),"center",bg_,"₹#,##0.00","0000FF",False),
        (row_data[6],"center",bg_,None,"000000",False),
        (rr_formula,"center",bg_,"0.0x",DARK_GREEN,True),
        (conf,"center",bg_,None,"000000",False),
        (float(row_data[9].replace(",","")) if row_data[9].replace(",","").isdigit() else 50000,"center",LIGHT_GOLD,"₹#,##0","0000FF",True),
        (pnl_t1,"center",LIGHT_GREEN,"₹#,##0;(₹#,##0);-",DARK_GREEN,True),
        (pnl_t2,"center",LIGHT_GREEN,"₹#,##0;(₹#,##0);-",DARK_GREEN,True),
        (stat,"center",stat_bg,None,"000000",True),
    ]
    for j,(v,ha,bg_use,fmt,fg_,bold_) in enumerate(entries):
        c=ws_sw.cell(row=r, column=2+j, value=v)
        c.fill=fill(bg_use); c.alignment=align(ha); c.border=border()
        c.font=Font(bold=bold_,color=fg_,size=9,name="Arial")
        if fmt: c.number_format=fmt
    ws_sw.row_dimensions[r].height=20

# Swing Trading Strategy Box
r_strat=13
section_title(ws_sw, r_strat, 2, "📘 SWING TRADING WITH BUYBACK DEEP VALUE — STRATEGY GUIDE", 12)
strats=[
    ("WHY BUYBACK STOCKS?","Company is buying back at higher price = floor support | Management confident in business | Less downside risk"),
    ("ENTRY TIMING","Enter when stock dips toward buyback accumulation zone | Confirm RSI below 45 | Volume drying up (sellers exhausted)"),
    ("POSITION SIZING","Never risk more than 2% of total capital per trade | Example: ₹5L portfolio → max ₹10,000 risk per swing trade"),
    ("STOP LOSS RULE","Always set GTT stop immediately after entry | Stop = 5-7% below entry OR below recent swing low"),
    ("TARGET SETTING","Target 1 = Previous resistance (take 50% profit) | Target 2 = Next major resistance (trail stop for rest)"),
    ("HOLDING PERIOD","Swing trades: 5-15 trading days typically | Exit at target even if you think it will go higher"),
    ("EXIT SIGNALS","Price hits Target 1 = book 50% | RSI crosses above 70 = start exiting | Volume dries up at resistance"),
    ("AVOID TRAPS","Never swing trade during budget/RBI/election event risk | Avoid stocks with pending court orders/investigations"),
]
for i,(cat,desc) in enumerate(strats):
    r=r_strat+1+i; ws_sw.row_dimensions[r].height=22
    c1=ws_sw.cell(row=r,column=2,value=cat)
    c1.font=Font(bold=True,color=WHITE,size=9,name="Arial"); c1.fill=fill(ORANGE if i%2==0 else DARK_NAVY)
    c1.alignment=align("center"); c1.border=border()
    ws_sw.merge_cells(f"C{r}:M{r}")
    c2=ws_sw.cell(row=r,column=3,value=desc)
    c2.font=font(size=9); c2.fill=fill(LIGHT_ORANGE if i%2==0 else WHITE)
    c2.alignment=align("left"); c2.border=border()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — CHARTS
# ══════════════════════════════════════════════════════════════════════════════
ws_ch = wb.create_sheet("📈 Charts & Analysis")
ws_ch.sheet_view.showGridLines = False
set_col_widths(ws_ch, {"A":2,"B":16,"C":12,"D":12,"E":12,"F":12,"G":12,"H":12,"I":12})

ws_ch.merge_cells("B1:I1")
c=ws_ch["B1"]; c.value="📈  PORTFOLIO CHARTS & VISUAL ANALYSIS"
c.font=Font(bold=True,color=WHITE,size=15,name="Arial"); c.fill=fill(DARK_NAVY)
c.alignment=align("center"); ws_ch.row_dimensions[1].height=36

# Data for Sector Allocation chart
section_title(ws_ch, 2, 2, "SECTOR ALLOCATION DATA (auto-linked to portfolio)", 7)
apply_header_row(ws_ch, 3, ["Sector","Invested (₹)","Current Value (₹)","P&L (₹)","% of Portfolio"], 2, BLUE_HEADER)
sectors=[("IT","3,50,000","4,05,000","55,000","32%"),("Banking","1,23,750","1,18,500","-5,250","12%"),
         ("Energy","1,20,000","1,26,000","6,000","10%"),("Pharma","72,000","74,000","2,000","7%"),
         ("Infra","96,000","1,04,000","8,000","10%"),("Utilities","72,000","78,000","6,000","8%"),
         ("Mining","36,750","38,000","1,250","4%"),("NBFC","2,04,000","2,20,000","16,000","17%")]
for i,row_data in enumerate(sectors):
    r=4+i; bg_=LIGHT_BLUE if i%2==0 else WHITE
    for j,v in enumerate(row_data):
        data_cell(ws_ch,r,2+j,v,bg_)

# Pie Chart — Sector Allocation
pie=PieChart()
pie.title="Portfolio Sector Allocation"
labels=Reference(ws_ch, min_col=2, min_row=4, max_row=11)
data=Reference(ws_ch, min_col=6, min_row=3, max_row=11)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.style=10
pie.width=14; pie.height=12
ws_ch.add_chart(pie, "B13")

# Monthly P&L data for Line Chart
section_title(ws_ch, 2, 9, "MONTHLY P&L TREND (Enter manually)", 2)
apply_header_row(ws_ch, 3, ["Month","P&L (₹)"], 9, BLUE_HEADER)
months=[("Jan-25","8500"),("Feb-25","12000"),("Mar-25","-3200"),("Apr-25","15000"),
        ("May-25","9800"),("Jun-25","18000"),("Jul-25","-5000"),("Aug-25","22000"),
        ("Sep-25","14000"),("Oct-25","19000"),("Nov-25","7500"),("Dec-25","25000")]
for i,(m,p) in enumerate(months):
    r=4+i; bg_=LIGHT_BLUE if i%2==0 else WHITE
    data_cell(ws_ch,r,9,m,bg_)
    c=data_cell(ws_ch,r,10,int(p),bg_,num_fmt="₹#,##0;(₹#,##0);-")
    c.font=Font(color=DARK_GREEN if int(p)>0 else DARK_RED, size=9, name="Arial")

# Line Chart — Monthly P&L
line=LineChart()
line.title="Monthly P&L Trend (₹)"
line.style=10; line.y_axis.title="P&L (₹)"; line.x_axis.title="Month"
line_data=Reference(ws_ch, min_col=10, min_row=3, max_row=15)
line.add_data(line_data, titles_from_data=True)
line_cats=Reference(ws_ch, min_col=9, min_row=4, max_row=15)
line.set_categories(line_cats)
line.width=16; line.height=12
ws_ch.add_chart(line, "H13")

# Risk Meter section
r_risk = 32
section_title(ws_ch, r_risk, 2, "🔔 PORTFOLIO RISK SCORECARD", 8)
apply_header_row(ws_ch, r_risk+1, ["Risk Factor","Your Score","Benchmark","Status","Comment"], 2, DARK_RED)
risk_items=[
    ("Diversification","8/10","7/10","✅ GOOD","Spread across 8+ sectors"),
    ("Stop Loss Coverage","6/10","9/10","⚠️ LOW","Set GTT for all open positions"),
    ("Leverage/Margin","10/10","10/10","✅ GOOD","No margin — cash delivery only"),
    ("Concentration Risk","7/10","8/10","⚠️ MEDIUM","IT sector slightly over-weighted"),
    ("Cash Reserve","5/10","8/10","🔴 LOW","Keep 15-20% cash for opportunities"),
    ("Overall Risk Score","7.2/10","8.4/10","⚠️ MEDIUM","Review stop losses and cash"),
]
for i,(factor,score,bench,status,comment) in enumerate(risk_items):
    r=r_risk+2+i; bg_=LIGHT_RED if "LOW" in status or "🔴" in status else (LIGHT_GOLD if "MEDIUM" in status else LIGHT_GREEN)
    for j,v in enumerate([factor,score,bench,status,comment]):
        c=data_cell(ws_ch,r,2+j,v,bg_,h_align="left" if j in [0,4] else "center")
    ws_ch.row_dimensions[r].height=18

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — WATCHLIST / NOTES
# ══════════════════════════════════════════════════════════════════════════════
ws_wl = wb.create_sheet("📝 Watchlist & Notes")
ws_wl.sheet_view.showGridLines = False
set_col_widths(ws_wl, {"A":2,"B":22,"C":12,"D":14,"E":14,"F":14,"G":16,"H":16,"I":20})

ws_wl.merge_cells("B1:I1")
c=ws_wl["B1"]; c.value="📝  WATCHLIST, RESEARCH NOTES & LEARNING TRACKER"
c.font=Font(bold=True,color=WHITE,size=15,name="Arial"); c.fill=fill(TEAL)
c.alignment=align("center"); ws_wl.row_dimensions[1].height=36

# Watchlist
section_title(ws_wl, 2, 2, "👁 WATCHLIST — STOCKS TO BUY ON NEXT DIP", 8)
apply_header_row(ws_wl, 3, ["Company","Sector","CMP","Target Entry","Watch Reason","Priority","Research Date","Notes"], 2, TEAL)
wl_data=[
    ("MARUTI SUZUKI","Auto","11500","10800-11000","EV transition + strong volumes","HIGH","",""),
    ("ASIAN PAINTS","Consumer","2600","2400-2500","Volume recovery expected Q4","HIGH","",""),
    ("DIVI'S LAB","Pharma","4800","4500-4700","API export growth story","MEDIUM","",""),
    ("PIDILITE","Consumer","2900","2700-2800","Rural demand recovery","MEDIUM","",""),
    ("HDFC LIFE","Insurance","680","640-660","Insurance penetration growing","HIGH","",""),
    ("ZOMATO","Tech","220","190-210","Food delivery market leader","LOW","",""),
    ("ADANIPORTS","Infra","1350","1250-1300","Cargo volume growth","MEDIUM","",""),
    ("MUTHOOT FIN","NBFC","1900","1750-1850","Gold loan demand rising","LOW","",""),
]
for i,row_data in enumerate(wl_data):
    r=4+i; bg_=LIGHT_TEAL if i%2==0 else WHITE
    pri=row_data[5]
    pri_bg=LIGHT_GREEN if pri=="HIGH" else (LIGHT_GOLD if pri=="MEDIUM" else LIGHT_GRAY)
    for j,v in enumerate(row_data):
        bg_use=pri_bg if j==5 else bg_
        c=data_cell(ws_wl,r,2+j,v,bg_use,h_align="left" if j in [0,4,7] else "center")
    ws_wl.row_dimensions[r].height=20

# My Trading Rules
r_rules=14
section_title(ws_wl, r_rules, 2, "📜 MY PERSONAL TRADING RULES (Customize These)", 8)
my_rules=[
    "1. NEVER invest more than 10% of portfolio in a single stock",
    "2. ALWAYS set stop loss / GTT within 24 hours of buying",
    "3. Buy in 2-3 tranches — never go all-in at once",
    "4. Minimum 1:2 Risk:Reward ratio for every trade",
    "5. Book 50% profit when first target is hit, no matter what",
    "6. No averaging down if fundamental story has changed",
    "7. Keep 15-20% cash always — opportunity fund",
    "8. Review portfolio every Sunday evening",
    "9. Don't trade on tips — only trade what you understand",
    "10. Log every trade here — review monthly to improve",
]
for i,rule in enumerate(my_rules):
    r=r_rules+1+i
    ws_wl.merge_cells(f"B{r}:I{r}")
    c=ws_wl.cell(row=r,column=2,value=rule)
    c.font=Font(bold=i%2==0,color=DARK_NAVY,size=9,name="Arial")
    c.fill=fill(LIGHT_TEAL if i%2==0 else WHITE); c.alignment=align("left"); c.border=border()
    ws_wl.row_dimensions[r].height=18

# Trade Journal
r_j=26
section_title(ws_wl, r_j, 2, "📒 TRADE JOURNAL — LOG EVERY TRADE", 8)
apply_header_row(ws_wl, r_j+1, ["Date","Stock","Buy Price","Sell Price","Qty","P&L (₹)","% Return","Lesson Learned"], 2, TEAL)
for i in range(10):
    r=r_j+2+i; bg_=LIGHT_TEAL if i%2==0 else WHITE
    for j in range(8):
        c=ws_wl.cell(row=r,column=2+j)
        c.fill=fill(bg_); c.alignment=align("center"); c.border=border()
        if j==5:
            c.value=f"=IF(D{r}=0,\"\",ROUND((D{r}-C{r})*E{r},2))"
            c.number_format="₹#,##0.00;(₹#,##0.00);-"
        elif j==6:
            c.value=f"=IF(C{r}=0,\"\",ROUND((D{r}-C{r})/C{r}*100,2))"
            c.number_format="0.00%;(0.00%);-"
    ws_wl.row_dimensions[r].height=18

# ── Tab colors ──
ws_dash.sheet_properties.tabColor = "1B2A4A"
ws_port.sheet_properties.tabColor = "1F4E79"
ws_buy.sheet_properties.tabColor  = "1A6B3C"
ws_bb.sheet_properties.tabColor   = "7030A0"
ws_gtt.sheet_properties.tabColor  = "C9A84C"
ws_sw.sheet_properties.tabColor   = "C55A11"
ws_ch.sheet_properties.tabColor   = "1B2A4A"
ws_wl.sheet_properties.tabColor   = "008080"

# Sheet order / freeze panes
for ws in [ws_dash, ws_port, ws_buy, ws_bb, ws_gtt, ws_sw, ws_wl]:
    ws.freeze_panes = "B3"

out = "/home/claude/Investor_Portfolio_Report_Template.xlsx"
wb.save(out)
print("Saved:", out)
