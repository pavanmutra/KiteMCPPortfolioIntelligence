import json
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

DARK_NAVY = "1B2A4A"
GOLD = "C9A84C"
LIGHT_GOLD = "F5E6C0"
DARK_GREEN = "1A6B3C"
LIGHT_GREEN = "D6F0E0"
DARK_RED = "8B0000"
LIGHT_RED = "FFE0E0"
BLUE_HEADER = "1F4E79"
LIGHT_BLUE = "DEEAF1"
ORANGE = "C55A11"
LIGHT_ORANGE = "FCE4D6"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
PURPLE = "7030A0"
LIGHT_PURPLE = "EAD1F5"

def side(style="thin", color="BFBFBF"):
    return Side(style=style, color=color)

def border(all="thin"):
    s = side(all)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = side("medium", "1B2A4A")
    return Border(left=t, right=t, top=t, bottom=t)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col_widths(ws, widths_dict):
    for col, w in widths_dict.items():
        ws.column_dimensions[col].width = w

def apply_header_row(ws, row, headers, start_col=1, bg=BLUE_HEADER, fg=WHITE):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col+i, value=h)
        c.font = font(bold=True, color=fg, size=9)
        c.fill = fill(bg)
        c.alignment = align("center")
        c.border = border("thin")

def section_title(ws, row, col, text, span=1, bg=DARK_NAVY, fg=GOLD):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, color=fg, size=11, name="Arial")
    c.fill = fill(bg)
    c.alignment = align("center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c.border = thick_border()

def data_cell(ws, row, col, value, bg=WHITE, fg="000000", bold=False, num_fmt=None, h_align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h_align)
    c.border = border()
    if num_fmt:
        c.number_format = num_fmt
    return c

wb = Workbook()

# Load live holdings data
holdings = [
    {"symbol": "ABB", "qty": 2, "avg": 6780, "cmp": 7148, "pnl": 736},
    {"symbol": "CAMS", "qty": 244, "avg": 707.892, "cmp": 749.25, "pnl": 10091.35},
    {"symbol": "COALINDIA", "qty": 300, "avg": 432.5, "cmp": 441.75, "pnl": 2775},
    {"symbol": "ENERGY", "qty": 2571, "avg": 36.081789, "cmp": 39.41, "pnl": 8556.83},
    {"symbol": "FMCGIETF", "qty": 2177, "avg": 49.212586, "cmp": 53.25, "pnl": 8789.45},
    {"symbol": "GICRE", "qty": 150, "avg": 400.05, "cmp": 395.9, "pnl": -622.5},
    {"symbol": "HINDALCO", "qty": 25, "avg": 1040, "cmp": 1015.25, "pnl": -618.75},
    {"symbol": "JINDALPHOT", "qty": 85, "avg": 1320.710588, "cmp": 1143.85, "pnl": -15033.15},
    {"symbol": "JUSTDIAL", "qty": 200, "avg": 504.42275, "cmp": 546.45, "pnl": 8405.45},
    {"symbol": "KNRCON", "qty": 500, "avg": 118, "cmp": 120.96, "pnl": 1480},
    {"symbol": "LICHSGFIN", "qty": 50, "avg": 530.4, "cmp": 536, "pnl": 280},
    {"symbol": "LICI", "qty": 30, "avg": 838, "cmp": 828.2, "pnl": -294},
    {"symbol": "LIQUIDETF", "qty": 100, "avg": 1000, "cmp": 1000, "pnl": 0},
    {"symbol": "NCC", "qty": 300, "avg": 160.36, "cmp": 160.45, "pnl": 27},
    {"symbol": "NMDC", "qty": 500, "avg": 85.97, "cmp": 88.81, "pnl": 1420},
    {"symbol": "NTPC", "qty": 160, "avg": 391.875, "cmp": 397.9, "pnl": 964},
    {"symbol": "NXST-RR", "qty": 650, "avg": 135.185938, "cmp": 157.81, "pnl": 14705.64},
    {"symbol": "POWERGRID", "qty": 100, "avg": 313, "cmp": 319.7, "pnl": 670},
    {"symbol": "RECLTD", "qty": 100, "avg": 365, "cmp": 382.2, "pnl": 1720},
    {"symbol": "RELIANCE", "qty": 35, "avg": 1335, "cmp": 1362.6, "pnl": 966},
    {"symbol": "SBIN", "qty": 75, "avg": 1068, "cmp": 1107.6, "pnl": 2970},
    {"symbol": "TMCV", "qty": 160, "avg": 367.132729, "cmp": 438.25, "pnl": 11378.76},
    {"symbol": "VHL", "qty": 45, "avg": 3545.7, "cmp": 3361, "pnl": -8311.5},
]

# Load GTT data
with open(r'C:\Users\pc\.local\share\opencode\tool-output\tool_dab92d785001I8a6dqnaLre9c7') as f:
    gtts = json.load(f)
active_gtts = [g for g in gtts if g['status'] == 'active']
print(f"Loaded {len(holdings)} holdings, {len(active_gtts)} active GTTs")

# SHEET 1 - DASHBOARD
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_view.showGridLines = False
ws_dash.row_dimensions[1].height = 50
set_col_widths(ws_dash, {"A":2,"B":16,"C":12,"D":12,"E":12,"F":12,"G":12,"H":12,"I":12,"J":12,"K":12,"L":12})

ws_dash.merge_cells("B1:L1")
c = ws_dash["B1"]
c.value = "INVESTOR PORTFOLIO COMMAND CENTER"
c.font = Font(bold=True, color=WHITE, size=20, name="Arial")
c.fill = fill(DARK_NAVY)
c.alignment = align("center")

total_value = sum(h['qty'] * h['cmp'] for h in holdings)
total_invested = sum(h['qty'] * h['avg'] for h in holdings)
total_pnl = sum(h['pnl'] for h in holdings)
total_pnl_pct = (total_pnl / total_invested * 100) if total_invested > 0 else 0

ws_dash.merge_cells("B2:L2")
c = ws_dash["B2"]
c.value = f"Report Date: {datetime.date.today().strftime('%d %B %Y')} | Holdings: {len(holdings)} | GTTs: {len(active_gtts)}"
c.font = font(color=GOLD, size=10, bold=True)
c.fill = fill(BLUE_HEADER)
c.alignment = align("center")
ws_dash.row_dimensions[2].height = 18

# KPI Cards
ws_dash.merge_cells("B3:L3")
c = ws_dash["B3"]
c.value = "KEY PORTFOLIO METRICS"
c.font = Font(bold=True, color=WHITE, size=10, name="Arial")
c.fill = fill(BLUE_HEADER)
c.alignment = align("center")
ws_dash.row_dimensions[3].height = 16

kpi_data = [
    ("Total Portfolio Value", f"₹{total_value:,.0f}"),
    ("Total Invested", f"₹{total_invested:,.0f}"),
    ("Overall P&L", f"₹{total_pnl:,.0f}"),
    ("P&L %", f"{total_pnl_pct:.1f}%"),
    ("GTT Orders Active", str(len(active_gtts))),
]
kpi_colors = [BLUE_HEADER, DARK_NAVY, DARK_GREEN if total_pnl > 0 else DARK_RED, GOLD, PURPLE]

for idx, (label, val) in enumerate(kpi_data):
    col = chr(66 + idx * 2)
    if idx < 4:
        ws_dash.merge_cells(f"{col}4:{chr(67 + idx * 2)}4")
        ws_dash.merge_cells(f"{col}5:{chr(67 + idx * 2)}5")
    c = ws_dash[f"{col}4"]
    c.value = label
    c.font = Font(bold=True, color=WHITE, size=9, name="Arial")
    c.fill = fill(kpi_colors[idx])
    c.alignment = align("center")
    c.border = border("medium")
    ws_dash.row_dimensions[4].height = 20
    
    v = ws_dash[f"{col}5"]
    v.value = val
    v.font = Font(bold=True, color=WHITE, size=14, name="Arial")
    v.fill = fill(kpi_colors[idx])
    v.alignment = align("center")
    v.border = border("medium")
    ws_dash.row_dimensions[5].height = 28

# Portfolio Summary mini-table
r = 8
section_title(ws_dash, r, 2, "PORTFOLIO SNAPSHOT", 5)
headers_snap = ["Stock", "Qty", "Avg Cost", "CMP", "Current Value", "P&L"]
apply_header_row(ws_dash, r+1, headers_snap, 2, BLUE_HEADER)
ws_dash.row_dimensions[r+1].height = 16

for i in range(min(10, len(holdings))):
    h = holdings[i]
    row = r+2+i
    bg_ = LIGHT_GRAY if i%2==0 else WHITE
    vals = [h['symbol'], h['qty'], f"₹{h['avg']:,.2f}", f"₹{h['cmp']:,.2f}", f"₹{h['qty']*h['cmp']:,.0f}", f"₹{h['pnl']:,.0f}"]
    for j, v in enumerate(vals):
        c = data_cell(ws_dash, row, 2+j, v, bg_, num_fmt="0" if j in [1] else "₹#,##0")
    # Formula for P&L color
    pl_cell = ws_dash.cell(row=row, column=7)
    if h['pnl'] > 0:
        pl_cell.fill = fill(LIGHT_GREEN)
    else:
        pl_cell.fill = fill(LIGHT_RED)

# SHEET 2 - PORTFOLIO
ws_port = wb.create_sheet("Portfolio")
ws_port.sheet_view.showGridLines = False
set_col_widths(ws_port, {"A":2,"B":20,"C":12,"D":12,"E":12,"F":14,"G":14,"H":14,"I":12,"J":14,"K":14,"L":14,"M":14})

ws_port.merge_cells("B1:N1")
c = ws_port["B1"]
c.value = "MY INVESTMENT PORTFOLIO"
c.font = Font(bold=True, color=WHITE, size=16, name="Arial")
c.fill = fill(DARK_NAVY)
c.alignment = align("center")
ws_port.row_dimensions[1].height = 36

headers_port = ["Stock / Company", "Qty", "Avg Buy Price (₹)", "CMP (₹)", "Total Invested (₹)", "Current Value (₹)", "P&L (₹)", "P&L %"]
apply_header_row(ws_port, 2, headers_port, 2, BLUE_HEADER)
ws_port.row_dimensions[2].height = 32

for i, h in enumerate(holdings):
    r = 3 + i
    bg_ = LIGHT_BLUE if i%2==0 else WHITE
    invested = h['qty'] * h['avg']
    current_val = h['qty'] * h['cmp']
    pnl_pct = (h['pnl'] / invested * 100) if invested > 0 else 0
    
    data_cell(ws_port, r, 2, h['symbol'], bg_, h_align="left")
    data_cell(ws_port, r, 3, h['qty'], bg_, num_fmt="0")
    data_cell(ws_port, r, 4, h['avg'], bg_, num_fmt="₹#,##0.00")
    data_cell(ws_port, r, 5, h['cmp'], bg_, num_fmt="₹#,##0.00")
    data_cell(ws_port, r, 6, invested, bg_, num_fmt="₹#,##0.00")
    data_cell(ws_port, r, 7, current_val, bg_, num_fmt="₹#,##0.00")
    pl_cell = data_cell(ws_port, r, 8, h['pnl'], bg_, num_fmt="₹#,##0.00;(₹#,##0.00);-")
    if h['pnl'] > 0:
        pl_cell.fill = fill(LIGHT_GREEN)
    else:
        pl_cell.fill = fill(LIGHT_RED)
    data_cell(ws_port, r, 9, f"{pnl_pct:.1f}%", bg_, num_fmt="0.00%")

# Totals Row
tr = 3 + len(holdings)
ws_port.cell(row=tr, column=2, value="TOTAL").font = Font(bold=True, size=10)
ws_port.cell(row=tr, column=2).fill=fill(DARK_NAVY)
ws_port.cell(row=tr, column=2).font=Font(bold=True, color=WHITE, size=10, name="Arial")
ws_port.cell(row=tr, column=2).alignment=align("center")
for c_idx, val in [(6, total_invested), (7, total_value), (8, total_pnl)]:
    c = ws_port.cell(row=tr, column=c_idx, value=val)
    c.font=Font(bold=True, size=10, name="Arial")
    c.fill=fill(LIGHT_GOLD)
    c.number_format="₹#,##0.00"
    c.alignment=align("center")
    c.border=thick_border()

# Note
note_row = tr + 2
ws_port.merge_cells(f"B{note_row}:N{note_row}")
c = ws_port.cell(row=note_row, column=2, value="Data from Kite Holdings - " + datetime.datetime.now().strftime("%d %b %Y %H:%M"))
c.font = Font(italic=True, color=DARK_NAVY, size=9, name="Arial")

# SHEET 3 - GTT PLAN
ws_gtt = wb.create_sheet("GTT Plan")
ws_gtt.sheet_view.showGridLines = False
set_col_widths(ws_gtt, {"A":2,"B":16,"C":12,"D":12,"E":12,"F":12,"G":12,"H":12,"I":12})

ws_gtt.merge_cells("B1:I1")
c = ws_gtt["B1"]
c.value = "GTT (GOOD TILL TRIGGERED) ORDER PLANNER"
c.font = Font(bold=True, color=WHITE, size=15, name="Arial")
c.fill = fill(DARK_NAVY)
c.alignment = align("center")
ws_gtt.row_dimensions[1].height = 36

ws_gtt.merge_cells("B2:I2")
c = ws_gtt["B2"]
c.value = f"Active GTT Orders: {len(active_gtts)} | Generated: {datetime.datetime.now().strftime('%d %b %Y')}"
c.font = Font(italic=True, color=DARK_NAVY, size=9, name="Arial")
c.fill = fill(LIGHT_GOLD)
c.alignment = align("center")

headers_gtt = ["Stock", "CMP (₹)", "Trigger (₹)", "Type", "Qty", "Order Value (₹)", "Broker"]
apply_header_row(ws_gtt, 3, headers_gtt, 2, DARK_NAVY)
ws_gtt.row_dimensions[3].height = 28

for i in range(min(20, len(active_gtts))):
    g = active_gtts[i]
    r = 4 + i
    bg_ = LIGHT_GRAY if i%2==0 else WHITE
    sym = g['condition']['tradingsymbol']
    cmp = g['condition']['last_price']
    trigger = g['condition']['trigger_values'][0]
    order = g['orders'][0]
    qty = order['quantity']
    order_val = qty * trigger
    
    data_cell(ws_gtt, r, 2, sym, bg_, h_align="left")
    data_cell(ws_gtt, r, 3, cmp, bg_, num_fmt="₹#,##0.00")
    data_cell(ws_gtt, r, 4, trigger, bg_, num_fmt="₹#,##0.00")
    tt = "BUY" if order['transaction_type'] == "BUY" else "SELL"
    type_bg = LIGHT_GREEN if tt == "BUY" else LIGHT_RED
    t_cell = data_cell(ws_gtt, r, 5, tt, type_bg)
    t_cell.font = Font(bold=True, color=DARK_GREEN if tt == "BUY" else DARK_RED)
    data_cell(ws_gtt, r, 6, qty, bg_, num_fmt="0")
    data_cell(ws_gtt, r, 7, order_val, bg_, num_fmt="₹#,##0.00")
    data_cell(ws_gtt, r, 8, "Kite", bg_)

out_path = r"G:\des\kitemcp\investor_template_package\Live_Portfolio_Report.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Total Value: ₹{total_value:,.0f}")
print(f"Total Invested: ₹{total_invested:,.0f}")
print(f"Total P&L: ₹{total_pnl:,.0f} ({total_pnl_pct:.1f}%)")