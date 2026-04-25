import json
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK_NAVY = "1B2A4A"
GOLD = "C9A84C"
LIGHT_GOLD = "F5E6C0"
DARK_GREEN = "1A6B3C"
LIGHT_GREEN = "D6F0E0"
DARK_RED = "8B0000"
LIGHT_RED = "FFE0E0"
BLUE_HEADER = "1F4E79"
LIGHT_BLUE = "DEEAF1"
ORANGE = "C55A11"
LIGHT_ORANGE = "FCE4D6"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"
PURPLE = "7030A0"
LIGHT_PURPLE = "EAD1F5"
TEAL = "008080"
LIGHT_TEAL = "D0EBEB"

def side(style="thin", color="BFBFBF"):
    return Side(style=style, color=color)

def border(all="thin"):
    s = side(all)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = side("medium", "1B2A4A")
    return Border(left=t, right=t, top=t, bottom=t)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def apply_header_row(ws, row, headers, start_col=2, bg=BLUE_HEADER, fg=WHITE):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col+i, value=h)
        c.font = font(bold=True, color=fg, size=9)
        c.fill = fill(bg)
        c.alignment = align("center")
        c.border = border("thin")

def section_title(ws, row, col, text, span=1, bg=DARK_NAVY, fg=GOLD):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, color=fg, size=11, name="Arial")
    c.fill = fill(bg)
    c.alignment = align("center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c.border = thick_border()

def data_cell(ws, row, col, value, bg=WHITE, fg="000000", bold=False, num_fmt=None, h_align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h_align)
    c.border = border()
    if num_fmt:
        c.number_format = num_fmt
    return c

wb = Workbook()

# Holdings data
holdings = [
    {"symbol": "ABB", "qty": 2, "avg": 6780, "cmp": 7148, "pnl": 736},
    {"symbol": "CAMS", "qty": 244, "avg": 707.892, "cmp": 749.25, "pnl": 10091.35},
    {"symbol": "COALINDIA", "qty": 300, "avg": 432.5, "cmp": 441.75, "pnl": 2775},
    {"symbol": "ENERGY", "qty": 2571, "avg": 36.081789, "cmp": 39.41, "pnl": 8556.83},
    {"symbol": "FMCGIETF", "qty": 2177, "avg": 49.212586, "cmp": 53.25, "pnl": 8789.45},
    {"symbol": "GICRE", "qty": 150, "avg": 400.05, "cmp": 395.9, "pnl": -622.5},
    {"symbol": "HINDALCO", "qty": 25, "avg": 1040, "cmp": 1015.25, "pnl": -618.75},
    {"symbol": "JINDALPHOT", "qty": 85, "avg": 1320.710588, "cmp": 1143.85, "pnl": -15033.15},
    {"symbol": "JUSTDIAL", "qty": 200, "avg": 504.42275, "cmp": 546.45, "pnl": 8405.45},
    {"symbol": "KNRCON", "qty": 500, "avg": 118, "cmp": 120.96, "pnl": 1480},
    {"symbol": "LICHSGFIN", "qty": 50, "avg": 530.4, "cmp": 536, "pnl": 280},
    {"symbol": "LICI", "qty": 30, "avg": 838, "cmp": 828.2, "pnl": -294},
    {"symbol": "LIQUIDETF", "qty": 100, "avg": 1000, "cmp": 1000, "pnl": 0},
    {"symbol": "NCC", "qty": 300, "avg": 160.36, "cmp": 160.45, "pnl": 27},
    {"symbol": "NMDC", "qty": 500, "avg": 85.97, "cmp": 88.81, "pnl": 1420},
    {"symbol": "NTPC", "qty": 160, "avg": 391.875, "cmp": 397.9, "pnl": 964},
    {"symbol": "NXST-RR", "qty": 650, "avg": 135.185938, "cmp": 157.81, "pnl": 14705.64},
    {"symbol": "POWERGRID", "qty": 100, "avg": 313, "cmp": 319.7, "pnl": 670},
    {"symbol": "RECLTD", "qty": 100, "avg": 365, "cmp": 382.2, "pnl": 1720},
    {"symbol": "RELIANCE", "qty": 35, "avg": 1335, "cmp": 1362.6, "pnl": 966},
    {"symbol": "SBIN", "qty": 75, "avg": 1068, "cmp": 1107.6, "pnl": 2970},
    {"symbol": "TMCV", "qty": 160, "avg": 367.132729, "cmp": 438.25, "pnl": 11378.76},
    {"symbol": "VHL", "qty": 45, "avg": 3545.7, "cmp": 3361, "pnl": -8311.5},
]

with open(r'C:\Users\pc\.local\share\opencode\tool-output\tool_dab92d785001I8a6dqnaLre9c7') as f:
    gtts = json.load(f)
active_gtts = [g for g in gtts if g['status'] == 'active']

total_value = sum(h['qty'] * h['cmp'] for h in holdings)
total_invested = sum(h['qty'] * h['avg'] for h in holdings)
total_pnl = sum(h['pnl'] for h in holdings)

print(f"Loaded {len(holdings)} holdings, {len(active_gtts)} GTTs")

# SHEET 1 - DASHBOARD
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_view.showGridLines = False
set_col_widths(ws_dash, {"A":2,"B":18,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14,"I":14,"J":14,"K":14,"L":14})

ws_dash.merge_cells("B1:L1")
c = ws_dash["B1"]
c.value = "INVESTOR PORTFOLIO COMMAND CENTER"
c.font = Font(bold=True, color=WHITE, size=20, name="Arial")
c.fill = fill(DARK_NAVY)
c.alignment = align("center")

ws_dash.merge_cells("B2:L2")
c = ws_dash["B2"]
c.value = f"Report Date: {datetime.date.today().strftime('%d %B %Y')}   |   Holdings: {len(holdings)} | GTTs: {len(active_gtts)}"
c.font = font(color=GOLD, size=10, bold=True)
c.fill = fill(BLUE_HEADER)
c.alignment = align("center")
ws_dash.row_dimensions[2].height = 18

ws_dash.merge_cells("B3:L3")
c = ws_dash["B3"]
c.value = "KEY PORTFOLIO METRICS"
c.font = Font(bold=True, color=WHITE, size=10, name="Arial")
c.fill = fill(BLUE_HEADER)
c.alignment = align("center")
ws_dash.row_dimensions[3].height = 16

kpi_data = [
    ("Total Portfolio Value", total_value, "₹#,##0.00", BLUE_HEADER, WHITE),
    ("Total Invested", total_invested, "₹#,##0.00", DARK_NAVY, WHITE),
    ("Overall P&L", total_pnl, "₹#,##0.00;(₹#,##0.00);-", DARK_GREEN, WHITE),
    ("P&L %", (total_pnl/total_invested*100) if total_invested else 0, "0.00%", GOLD, DARK_NAVY),
    ("GTT Orders Active", len(active_gtts), "0", PURPLE, WHITE),
]
kpi_positions = [("B","C"), ("D","E"), ("F","G"), ("H","I"), ("J","K")]

for idx, (label, val, fmt, bg, fg) in enumerate(kpi_data):
    c1, c2 = kpi_positions[idx]
    ws_dash.merge_cells(f"{c1}4:{c2}4")
    ws_dash.merge_cells(f"{c1}5:{c2}5")
    c = ws_dash[f"{c1}4"]
    c.value = label
    c.font = Font(bold=True, color=fg, size=9, name="Arial")
    c.fill = fill(bg)
    c.alignment = align("center")
    c.border = border("medium")
    ws_dash.row_dimensions[4].height = 20
    v = ws_dash[f"{c1}5"]
    v.value = val
    v.font = Font(bold=True, color=fg, size=14, name="Arial")
    v.fill = fill(bg)
    v.alignment = align("center")
    v.number_format = fmt
    v.border = border("medium")
    ws_dash.row_dimensions[5].height = 28

# Portfolio Snapshot mini-table
r = 8
section_title(ws_dash, r, 2, "PORTFOLIO SNAPSHOT", 5)
headers_snap = ["Stock", "Qty", "Avg Cost", "CMP", "Current Value", "P&L"]
apply_header_row(ws_dash, r+1, headers_snap, 2, BLUE_HEADER)
ws_dash.row_dimensions[r+1].height = 16

for i in range(min(10, len(holdings))):
    h = holdings[i]
    row = r+2+i
    bg_ = LIGHT_GRAY if i%2==0 else WHITE
    data_cell(ws_dash, row, 2, h['symbol'], bg_, h_align="left")
    data_cell(ws_dash, row, 3, h['qty'], bg_, num_fmt="0")
    data_cell(ws_dash, row, 4, h['avg'], bg_, num_fmt="₹#,##0.00")
    data_cell(ws_dash, row, 5, h['cmp'], bg_, num_fmt="₹#,##0.00")
    data_cell(ws_dash, row, 6, h['qty']*h['cmp'], bg_, num_fmt="₹#,##0.00")
    pl_cell = data_cell(ws_dash, row, 7, h['pnl'], bg_, num_fmt="₹#,##0.00;(₹#,##0.00);-")
    pl_cell.fill = fill(LIGHT_GREEN if h['pnl'] > 0 else LIGHT_RED)

# Buy Suggestions summary (right side)
section_title(ws_dash, 8, 8, "TOP BUY SUGGESTIONS", 4)
apply_header_row(ws_dash, 9, ["Stock", "Target", "Upside%"], 8, DARK_GREEN)

buys = [("RECLTD", "450", "18%"), ("NMDC", "110", "24%"), ("JUSTDIAL", "600", "10%"), ("NXST-RR", "180", "14%"), ("TMCV", "500", "14%")]
for i, (st, tgt, up) in enumerate(buys):
    bg_ = LIGHT_GREEN if i%2==0 else WHITE
    for j, v in enumerate([st, tgt, up]):
        data_cell(ws_dash, 10+i, 8+j, v, bg_)

# GTT mini
r2 = 16
section_title(ws_dash, r2, 2, "ACTIVE GTT ORDERS", 5)
apply_header_row(ws_dash, r2+1, ["Stock", "Trigger", "Type", "Status"], 2, PURPLE)
gtts_sample = active_gtts[:5]
for i, g in enumerate(gtts_sample):
    bg_ = LIGHT_PURPLE if i%2==0 else WHITE
    sym = g['condition']['tradingsymbol']
    trigger = g['condition']['trigger_values'][0]
    tt = g['orders'][0]['transaction_type']
    data_cell(ws_dash, r2+2+i, 2, sym, bg_, h_align="left")
    data_cell(ws_dash, r2+2+i, 3, trigger, bg_, num_fmt="₹#,##0.00")
    t_cell = data_cell(ws_dash, r2+2+i, 4, tt, bg_)
    t_cell.font = Font(bold=True, color=DARK_GREEN if tt=="BUY" else DARK_RED)
    data_cell(ws_dash, r2+2+i, 5, "Active", bg_)

# SHEET 2 - PORTFOLIO
ws_port = wb.create_sheet("Portfolio")
ws_port.sheet_view.showGridLines = False
set_col_widths(ws_port, {"A":2,"B":20,"C":12,"D":12,"E":12,"F":14,"G":16,"H":16,"I":14,"J":12,"K":14,"L":14,"M":14})

ws_port.merge_cells("B1:M1")
c = ws_port["B1"]
c.value = "MY INVESTMENT PORTFOLIO"
c.font = Font(bold=True, color=WHITE, size=16, name="Arial")
c.fill = fill(DARK_NAVY)
c.alignment = align("center")
ws_port.row_dimensions[1].height = 36

headers_port = ["Stock / Company", "Qty", "Avg Buy Price (₹)", "CMP (₹)", "Total Invested (₹)", "Current Value (₹)", "P&L (₹)", "P&L %", "Weight %"]
apply_header_row(ws_port, 2, headers_port, 2, BLUE_HEADER)
ws_port.row_dimensions[2].height = 32

for i, h in enumerate(holdings):
    r = 3 + i
    bg_ = LIGHT_BLUE if i%2==0 else WHITE
    invested = h['qty'] * h['avg']
    current_val = h['qty'] * h['cmp']
    pnl_pct = (h['pnl'] / invested * 100) if invested > 0 else 0
    weight = (current_val / total_value * 100) if total_value > 0 else 0
    
    data_cell(ws_port, r, 2, h['symbol'], bg_, h_align="left")
    data_cell(ws_port, r, 3, h['qty'], bg_, num_fmt="0")
    data_cell(ws_port, r, 4, h['avg'], bg_, num_fmt="₹#,##0.00")
    data_cell(ws_port, r, 5, h['cmp'], bg_, num_fmt="₹#,##0.00")
    data_cell(ws_port, r, 6, invested, bg_, num_fmt="₹#,##0.00")
    data_cell(ws_port, r, 7, current_val, bg_, num_fmt="₹#,##0.00")
    pl_cell = data_cell(ws_port, r, 8, h['pnl'], bg_, num_fmt="₹#,##0.00;(₹#,##0.00);-")
    pl_cell.fill = fill(LIGHT_GREEN if h['pnl'] > 0 else LIGHT_RED)
    data_cell(ws_port, r, 9, f"{pnl_pct:.1f}%", bg_, num_fmt="0.00%")
    data_cell(ws_port, r, 10, f"{weight:.1f}%", bg_, num_fmt="0.00%")

# Totals Row
tr = 3 + len(holdings)
ws_port.cell(row=tr, column=2, value="TOTAL")
ws_port.cell(row=tr, column=2).font = Font(bold=True, size=10)
ws_port.cell(row=tr, column=2).fill = fill(DARK_NAVY)
ws_port.cell(row=tr, column=2).font = Font(bold=True, color=WHITE, size=10, name="Arial")
ws_port.cell(row=tr, column=2).alignment = align("center")

for c_idx, val in [(6, total_invested), (7, total_value), (8, total_pnl)]:
    c = ws_port.cell(row=tr, column=c_idx, value=val)
    c.font = Font(bold=True, size=10, name="Arial")
    c.fill = fill(LIGHT_GOLD)
    c.number_format = "₹#,##0.00"
    c.alignment = align("center")
    c.border = thick_border()

# SHEET 3 - BUY SUGGESTIONS
ws_buy = wb.create_sheet("Buy Suggestions")
ws_buy.sheet_view.showGridLines = False
set_col_widths(ws_buy, {"A":2,"B":20,"C":12,"D":12,"E":12,"F":14,"G":14,"H":14,"I":16,"J":16,"K":16,"L":16})

ws_buy.merge_cells("B1:L1")
c = ws_buy["B1"]
c.value = "BUY SUGGESTIONS - DEEP VALUE"
c.font = Font(bold=True, color=WHITE, size=15, name="Arial")
c.fill = fill(DARK_GREEN)
c.alignment = align("center")
ws_buy.row_dimensions[1].height = 36

apply_header_row(ws_buy, 2, ["Company", "CMP (₹)", "Buy Zone", "Target", "Stop Loss", "Upside %", "Rating"], 2, DARK_GREEN)
ws_buy.row_dimensions[2].height = 30

# Deep value stocks from portfolio
deep_value = [
    ("RECLTD", 382.2, "340-360", "450", "340", "18%", "***"),
    ("NMDC", 88.81, "80-85", "110", "80", "24%", "***"),
    ("JUSTDIAL", 546.45, "500-520", "600", "500", "10%", "**"),
    ("NXST-RR", 157.81, "140-150", "180", "140", "14%", "***"),
    ("TMCV", 438.25, "400-420", "500", "400", "14%", "***"),
]
for i, row in enumerate(deep_value):
    r = 3 + i
    bg_ = LIGHT_GREEN if i%2==0 else WHITE
    for j, v in enumerate(row):
        data_cell(ws_buy, r, 2+j, v, bg_, h_align="left" if j==0 else "center")

# SHEET 4 - BUYBACK
ws_bb = wb.create_sheet("Buyback Opps")
ws_bb.sheet_view.showGridLines = False
set_col_widths(ws_bb, {"A":2,"B":22,"C":12,"D":14,"E":14,"F":12,"G":14})

ws_bb.merge_cells("B1:G1")
c = ws_bb["B1"]
c.value = "SHARE BUYBACK OPPORTUNITIES"
c.font = Font(bold=True, color=WHITE, size=15, name="Arial")
c.fill = fill(PURPLE)
c.alignment = align("center")
ws_bb.row_dimensions[1].height = 36

apply_header_row(ws_bb, 2, ["Company", "CMP (₹)", "Buyback Price", "Premium %", "Status", "Action"], 2, PURPLE)
ws_bb.row_dimensions[2].height = 30

# SHEET 5 - GTT PLAN
ws_gtt = wb.create_sheet("GTT Plan")
ws_gtt.sheet_view.showGridLines = False
set_col_widths(ws_gtt, {"A":2,"B":20,"C":12,"D":14,"E":14,"F":14,"G":14,"H":14,"I":14,"J":14,"K":20})

ws_gtt.merge_cells("B1:K1")
c = ws_gtt["B1"]
c.value = "GTT ORDER PLANNER"
c.font = Font(bold=True, color=WHITE, size=15, name="Arial")
c.fill = fill(DARK_NAVY)
c.alignment = align("center")
ws_gtt.row_dimensions[1].height = 36

ws_gtt.merge_cells("B2:K2")
c = ws_gtt["B2"]
c.value = f"Active GTT Orders: {len(active_gtts)}"
c.font = Font(italic=True, color=DARK_NAVY, size=9, name="Arial")
c.fill = fill(LIGHT_GOLD)
c.alignment = align("center")

apply_header_row(ws_gtt, 3, ["Stock", "CMP (₹)", "Trigger (₹)", "Type", "Qty", "Order Value (₹)", "Stop Loss", "Target", "R:R", "Broker", "Status"], 2, DARK_NAVY)
ws_gtt.row_dimensions[3].height = 28

for i in range(min(20, len(active_gtts))):
    g = active_gtts[i]
    r = 4 + i
    bg_ = LIGHT_GRAY if i%2==0 else WHITE
    sym = g['condition']['tradingsymbol']
    cmp = g['condition']['last_price']
    trigger = g['condition']['trigger_values'][0]
    order = g['orders'][0]
    qty = order['quantity']
    tt = order['transaction_type']
    type_bg = LIGHT_GREEN if tt == "BUY" else LIGHT_RED
    
    data_cell(ws_gtt, r, 2, sym, bg_, h_align="left")
    data_cell(ws_gtt, r, 3, cmp, bg_, num_fmt="₹#,##0.00")
    data_cell(ws_gtt, r, 4, trigger, bg_, num_fmt="₹#,##0.00")
    t_cell = data_cell(ws_gtt, r, 5, tt, type_bg)
    t_cell.font = Font(bold=True, color=DARK_GREEN if tt=="BUY" else DARK_RED)
    data_cell(ws_gtt, r, 6, qty, bg_, num_fmt="0")
    data_cell(ws_gtt, r, 7, qty*trigger, bg_, num_fmt="₹#,##0.00")
    data_cell(ws_gtt, r, 8, "", bg_, num_fmt="₹#,##0.00")
    data_cell(ws_gtt, r, 9, "", bg_, num_fmt="0.0x")
    data_cell(ws_gtt, r, 10, "Kite", bg_)
    data_cell(ws_gtt, r, 11, "Active", bg_)

# SHEET 6 - SWING TRADING
ws_sw = wb.create_sheet("Swing Trading")
ws_sw.sheet_view.showGridLines = False
set_col_widths(ws_sw, {"A":2,"B":20,"C":13,"D":12,"E":14,"F":14,"G":14,"H":12,"I":12,"J":12,"K":12,"L":16})

ws_sw.merge_cells("B1:L1")
c = ws_sw["B1"]
c.value = "SWING TRADING - BUYBACK OPPORTUNITIES"
c.font = Font(bold=True, color=WHITE, size=15, name="Arial")
c.fill = fill(ORANGE)
c.alignment = align("center")
ws_sw.row_dimensions[1].height = 36

apply_header_row(ws_sw, 2, ["Stock", "Setup", "Entry Zone", "Target 1", "Target 2", "Stop Loss", "Days", "R:R", "Confidence", "Capital", "P&L T1", "Status"], 2, ORANGE)
ws_sw.row_dimensions[2].height = 30

sw_data = [
    ("RECLTD", "Deep Value", "340-360", "400", "450", "320", "5-10", "2.0x", "High", "50000", "", "WATCH"),
    ("NMDC", "Buyback", "80-85", "95", "110", "75", "5-10", "2.5x", "High", "40000", "", "WATCH"),
    ("JUSTDIAL", "Recovery", "500-520", "560", "600", "480", "7-14", "2.0x", "Medium", "50000", "", "WATCH"),
]
for i, row in enumerate(sw_data):
    r = 3 + i
    bg_ = LIGHT_ORANGE if i%2==0 else WHITE
    for j, v in enumerate(row):
        data_cell(ws_sw, r, 2+j, v, bg_, h_align="left" if j==0 else "center")

# SHEET 7 - CHARTS
ws_ch = wb.create_sheet("Charts")
ws_ch.sheet_view.showGridLines = False
set_col_widths(ws_ch, {"A":2,"B":16,"C":12,"D":12,"E":12,"F":12,"G":12,"H":12,"I":12})

ws_ch.merge_cells("B1:I1")
c = ws_ch["B1"]
c.value = "PORTFOLIO CHARTS & ANALYSIS"
c.font = Font(bold=True, color=WHITE, size=15, name="Arial")
c.fill = fill(DARK_NAVY)
c.alignment = align("center")
ws_ch.row_dimensions[1].height = 36

section_title(ws_ch, 2, 2, "SECTOR ALLOCATION", 7)
apply_header_row(ws_ch, 3, ["Sector", "Value (₹)", "%"], 2, BLUE_HEADER)

# SHEET 8 - NOTES
ws_wl = wb.create_sheet("Watchlist")
ws_wl.sheet_view.showGridLines = False
set_col_widths(ws_wl, {"A":2,"B":22,"C":12,"D":14,"E":14,"F":14,"G":16})

ws_wl.merge_cells("B1:G1")
c = ws_wl["B1"]
c.value = "WATCHLIST & NOTES"
c.font = Font(bold=True, color=WHITE, size=15, name="Arial")
c.fill = fill(TEAL)
c.alignment = align("center")
ws_wl.row_dimensions[1].height = 36

section_title(ws_wl, 2, 2, "WATCHLIST - TO BUY ON DIP", 6)
apply_header_row(ws_wl, 3, ["Company", "CMP", "Target Entry", "Priority", "Reason"], 2, TEAL)
ws_wl.row_dimensions[3].height = 24

# Tab colors
ws_dash.sheet_properties.tabColor = "1B2A4A"
ws_port.sheet_properties.tabColor = "1F4E79"
ws_buy.sheet_properties.tabColor = "1A6B3C"
ws_bb.sheet_properties.tabColor = "7030A0"
ws_gtt.sheet_properties.tabColor = "C9A84C"
ws_sw.sheet_properties.tabColor = "C55A11"
ws_ch.sheet_properties.tabColor = "1B2A4A"
ws_wl.sheet_properties.tabColor = "008080"

# Freeze panes
for ws in [ws_dash, ws_port, ws_buy, ws_bb, ws_gtt, ws_sw, ws_wl]:
    ws.freeze_panes = "B3"

out = r"G:\des\kitemcp\investor_template_package\Live_Portfolio_Report.xlsx"
wb.save(out)
print(f"Saved: {out}")